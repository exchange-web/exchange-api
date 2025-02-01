from rest_framework import generics
from rest_framework.response import Response
from rest_framework.decorators import api_view
from data_entries.models import DataEntry
from .serializers import SummaryReportSerializer, ClientReportSerializer, OperationTypeReportSerializer
from django.db import models
from datetime import datetime

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

@api_view(['GET'])
def summary_report(request):
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')


    data_entries = DataEntry.objects.filter(date_created__range=[start_date, end_date])
    
    total_amount = data_entries.aggregate(total=models.Sum('amount'))['total']
    average_rate = data_entries.aggregate(avg=models.Avg('cross_rate'))['avg']
    
    report_data = {
        'total_amount': total_amount,
        'average_rate': average_rate,
        'client': 'All Clients',
        'currency_pair': 'All Pairs'
    }
    
    serializer = SummaryReportSerializer(report_data)
    return Response(serializer.data)

@api_view(['GET'])
def client_report(request, client_id):
    data_entries = DataEntry.objects.filter(client_id=client_id)
    serializer = ClientReportSerializer(data_entries, many=True)
    return Response(serializer.data)

@api_view(['GET'])
def operation_type_report(request):
    transaction_type = request.query_params.get('transaction_type')
    currency_pair = request.query_params.get('currency_pair')
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    
    data_entries = DataEntry.objects.filter(
        transaction_type=transaction_type,
        currency_pair=currency_pair,
        date__range=[start_date, end_date]
    )
    
    total_amount = data_entries.aggregate(total=models.Sum('amount'))['total']
    average_rate = data_entries.aggregate(avg=models.Avg('cross_rate'))['avg']
    
    report_data = {
        'total_amount': total_amount,
        'average_rate': average_rate,
        'transaction_type': transaction_type,
        'currency_pair': currency_pair
    }
    
    serializer = OperationTypeReportSerializer(report_data)
    return Response(serializer.data)

from django.http import HttpResponse
from openpyxl import load_workbook
from rest_framework.decorators import api_view
from data_entries.models import DataEntry
from datetime import datetime
from django.db.models import Sum, Avg

@api_view(['GET'])
def generate_report_xlsx(request):
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')

    # Load the template
    template_path = 'path/to/report_template.xlsx'
    wb = load_workbook(template_path)

    # Get the sheets
    ws_income = wb['приход']
    ws_expense = wb['расход']

    # Filter data entries
    data_entries = DataEntry.objects.filter(date_created__range=[start_date, end_date])

    # Populate the "приход" sheet
    row_num = 8
    total_income = 0
    total_exchange_rate_income = 0
    count_income = 0

    for entry in data_entries.filter(transaction_type='income'):
        date_str = entry.date_created.strftime("%d %b %Y")
        amount_in_currency = entry.amount_in
        exchange_rate = entry.cross_rate
        amount_converted = amount_in_currency * exchange_rate

        ws_income[f'A{row_num}'] = date_str
        ws_income[f'B{row_num}'] = entry.id
        ws_income[f'C{row_num}'] = entry.client.name
        ws_income[f'E{row_num}'] = amount_in_currency
        ws_income[f'F{row_num}'] = exchange_rate
        ws_income[f'G{row_num}'] = amount_converted

        total_income += amount_in_currency
        total_exchange_rate_income += exchange_rate
        count_income += 1
        row_num += 1

    if count_income > 0:
        average_rate_income = total_exchange_rate_income / count_income
    else:
        average_rate_income = 0

    ws_income['E3'] = total_income
    ws_income['F3'] = average_rate_income
    ws_income['G3'] = total_income * average_rate_income

    # Populate the "расход" sheet
    row_num = 8
    total_expense = 0
    total_exchange_rate_expense = 0
    count_expense = 0

    for entry in data_entries.filter(transaction_type='expense'):
        date_str = entry.date_created.strftime("%d %b %Y")
        amount_in_currency = entry.amount_in
        exchange_rate = entry.cross_rate
        amount_converted = amount_in_currency * exchange_rate

        ws_expense[f'A{row_num}'] = date_str
        ws_expense[f'B{row_num}'] = entry.id
        ws_expense[f'C{row_num}'] = entry.client.name
        ws_expense[f'E{row_num}'] = amount_in_currency
        ws_expense[f'F{row_num}'] = exchange_rate
        ws_expense[f'G{row_num}'] = amount_converted

        total_expense += amount_in_currency
        total_exchange_rate_expense += exchange_rate
        count_expense += 1
        row_num += 1

    if count_expense > 0:
        average_rate_expense = total_exchange_rate_expense / count_expense
    else:
        average_rate_expense = 0

    ws_expense['E3'] = total_expense
    ws_expense['F3'] = average_rate_expense
    ws_expense['G3'] = total_expense * average_rate_expense

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=report.xlsx'
    wb.save(response)

    return response


'''@api_view(['GET'])
def generate_report_xlsx(request):
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')

    data_entries = DataEntry.objects.filter(date_created__range=[start_date, end_date])

    wb = Workbook()
    ws = wb.active
    ws.title = "Data Entries Report"

    headers = ["Date", "ID", "Client", "Amount (Currency In)", "Exchange Rate", "Amount (Currency Out)"]
    ws.append(headers)

    for entry in data_entries:
        date_str = entry.date_created.strftime("%d %b %Y")
        amount_in_currency = entry.amount_in
        exchange_rate = entry.cross_rate
        amount_converted = amount_in_currency * exchange_rate
        row = [
            date_str,
            entry.id,
            entry.client.name,
            amount_in_currency,
            exchange_rate,
            amount_converted
        ]
        ws.append(row)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter 
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=report.xlsx'
    wb.save(response)

    return response'''